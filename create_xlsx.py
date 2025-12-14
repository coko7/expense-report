import sys
import csv
from openpyxl import load_workbook

xlsx_file = sys.argv[1]
csv_file = sys.argv[2]
employee_name = sys.argv[3]
employee_num = sys.argv[4]
expenses_date = sys.argv[5]

wb = load_workbook(xlsx_file)
sheet = wb['Blad1']

with open(csv_file, 'r') as f:
    rows = list(csv.reader(f))

sheet.cell(row=5, column=2).value = employee_name
sheet.cell(row=6, column=2).value = employee_num
sheet.cell(row=7, column=2).value = expenses_date

# Insert starting at row 10 (A10 = column 1, row 10)
for i, row in enumerate(rows):
    for j, value in enumerate(row):
        cell = sheet.cell(row=10+i, column=1+j)
        try:
            cell.value = float(value)
        except:
            cell.value = value

wb.save(xlsx_file)
print("✓ Data inserted into work sheet")
